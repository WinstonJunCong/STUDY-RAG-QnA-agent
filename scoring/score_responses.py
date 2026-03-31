#!/usr/bin/env python3
"""
LLM-Judged Scoring for RAG Responses

Scores all RAG response versions using Ollama LLM against ground truth.
Adds scores to UAT.xlsx and creates a Summary sheet.
"""

import json
import sys
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, str(Path(__file__).parent.parent))

from llama_index.core import Settings
from llama_index.llms.ollama import Ollama
from llama_index.embeddings.huggingface import HuggingFaceEmbedding
import config


SCORING_PROMPT = """You are evaluating a RAG system response against expected answer.

## Question
{question}

## Expected Answer
{expected_answer}

## Pass Condition
{pass_condition}

## System Response
{response}

## Scoring Rubric
Score 100: Perfect - matches expected answer exactly, all criteria met
Score 75-99: Good - captures key information, minor gaps
Score 50-74: Partial - captures some information, missing important parts
Score 0-49: Fail - incorrect or missing key information

## Special Cases
- Q9 (SLA Conflict): Must state BOTH 1 hour AND 30 minutes with conflict flag. Only 1 value = max 50.
- Q11 (Setup Steps): Must include all steps, NOT repeat from multiple sources
- Not Found questions (Q14, Q15): Must explicitly say not found in documents
- Yes/No questions: Must follow correct framing (No as first word when answer is No)

## Output (JSON only, no other text)
{{"score": 0-100, "rating": "Pass|Partial|Fail", "reasoning": "1-2 sentence explanation"}}
"""


def configure_llm():
    """Configure Ollama LLM for scoring."""
    Settings.embed_model = HuggingFaceEmbedding(model_name=config.EMBED_MODEL)
    Settings.llm = Ollama(
        model=config.OLLAMA_MODEL,
        base_url=config.OLLAMA_BASE_URL,
        request_timeout=120.0,
    )
    print(f"[scoring] LLM: {config.OLLAMA_MODEL}")


def get_response_columns(df):
    """Get all response columns from UAT dataframe."""
    response_cols = []
    for col in df.columns:
        if 'Latest Response' in col or col.startswith('Response'):
            # Extract version name
            match = re.search(r'\((R\d+[^)]*)\)', col)
            if match:
                version = match.group(1)
            elif col.startswith('Response ') and col != 'Response 6':
                version = col.replace('Response ', 'R')
            else:
                version = col
            response_cols.append((col, version))
    return response_cols


def score_response(question, expected, pass_condition, response):
    """Score a single response using LLM judge."""
    if pd.isna(response) or not response or str(response).strip() == '':
        return {'score': 0, 'rating': 'Fail', 'reasoning': 'Empty response'}
    
    prompt = SCORING_PROMPT.format(
        question=str(question),
        expected_answer=str(expected) if pd.notna(expected) else 'Not specified',
        pass_condition=str(pass_condition) if pd.notna(pass_condition) else 'None',
        response=str(response)
    )
    
    try:
        response_llm = Settings.llm.complete(prompt)
        result_text = str(response_llm).strip()
        
        # Parse JSON from response
        json_match = re.search(r'\{[^}]+\}', result_text, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group())
            return {
                'score': max(0, min(100, int(result.get('score', 0)))),
                'rating': result.get('rating', 'Fail'),
                'reasoning': result.get('reasoning', 'No reasoning provided')
            }
        else:
            return {'score': 0, 'rating': 'Fail', 'reasoning': f'Could not parse LLM response: {result_text[:100]}'}
    except Exception as e:
        return {'score': 0, 'rating': 'Fail', 'reasoning': f'Error: {str(e)[:50]}'}


def score_version(df, response_col, version_name, verbose=True):
    """Score all questions for a single version."""
    if verbose:
        print(f"\n  Scoring {version_name}...")
    
    scores = []
    ratings = []
    reasonings = []
    
    for idx, row in df.iterrows():
        question = row['Questiion']
        expected = row['Expected Answer']
        pass_cond = row['Pass Condition']
        response = row[response_col]
        
        result = score_response(question, expected, pass_cond, response)
        scores.append(result['score'])
        ratings.append(result['rating'])
        reasonings.append(result['reasoning'])
        
        if verbose:
            print(f"    Q{idx+1}: {result['score']} ({result['rating']})")
    
    # Calculate stats
    pass_count = sum(1 for r in ratings if r == 'Pass')
    partial_count = sum(1 for r in ratings if r == 'Partial')
    fail_count = sum(1 for r in ratings if r == 'Fail')
    avg_score = sum(scores) / len(scores) if scores else 0
    
    if verbose:
        print(f"  {version_name}: Pass={pass_count}, Partial={partial_count}, Fail={fail_count}, Avg={avg_score:.1f}%")
    
    return {
        'scores': scores,
        'ratings': ratings,
        'reasonings': reasonings,
        'stats': {
            'pass': pass_count,
            'partial': partial_count,
            'fail': fail_count,
            'avg': avg_score
        }
    }


def create_summary_sheet(wb, all_results):
    """Create Summary sheet with aggregate scores."""
    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    
    ws = wb.create_sheet('Score Summary')
    
    # Header
    headers = ['Version', 'Pass', 'Partial', 'Fail', 'Avg Score']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, (version, stats) in enumerate(all_results.items(), 2):
        ws.cell(row=row_idx, column=1, value=version)
        ws.cell(row=row_idx, column=2, value=stats['pass'])
        ws.cell(row=row_idx, column=3, value=stats['partial'])
        ws.cell(row=row_idx, column=4, value=stats['fail'])
        ws.cell(row=row_idx, column=5, value=f"{stats['avg']:.1f}%")
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    
    return ws


def format_uath_sheet(wb):
    """Format the UAT Responses sheet for readability."""
    ws = wb['UAT Responses'] if 'UAT Responses' in wb.sheetnames else wb.active
    
    # Set column widths and wrap text
    column_widths = {
        'A': 50,  # Questiion
        'B': 60,  # Expected Answer
        'C': 40,  # Pass Condition
    }
    
    for col in ws.columns:
        col_letter = col[0].column_letter
        width = column_widths.get(col_letter, 80)
        ws.column_dimensions[col_letter].width = width
        
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    return ws


def add_scores_to_excel(df, all_results, output_path):
    """Add scores to Excel file and create Summary sheet."""
    # Load workbook
    wb = load_workbook(output_path)
    
    # Check if we need to add a sheet or use existing
    sheet_name = 'UAT Responses'
    if sheet_name not in wb.sheetnames:
        # Create from dataframe
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb[sheet_name]
    
    # Add score columns for each version
    col_offset = len(df.columns) + 1
    
    for version, data in all_results.items():
        # Score column
        score_col = col_offset
        ws.cell(row=1, column=score_col, value=f'Score_{version}')
        ws.cell(row=1, column=score_col).font = Font(bold=True)
        ws.cell(row=1, column=score_col).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws.cell(row=1, column=score_col).font = Font(bold=True, color='FFFFFF')
        
        for row_idx, score in enumerate(data['scores'], 2):
            ws.cell(row=row_idx, column=score_col, value=score)
            ws.cell(row=row_idx, column=score_col).alignment = Alignment(horizontal='center')
        
        col_offset += 1
        
        # Rating column
        ws.cell(row=1, column=col_offset, value=f'Rating_{version}')
        ws.cell(row=1, column=col_offset).font = Font(bold=True)
        ws.cell(row=1, column=col_offset).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        ws.cell(row=1, column=col_offset).font = Font(bold=True, color='000000')
        
        for row_idx, rating in enumerate(data['ratings'], 2):
            ws.cell(row=row_idx, column=col_offset, value=rating)
            ws.cell(row=row_idx, column=col_offset).alignment = Alignment(horizontal='center')
        
        col_offset += 1
        
        # Reasoning column
        ws.cell(row=1, column=col_offset, value=f'Reasoning_{version}')
        ws.cell(row=1, column=col_offset).font = Font(bold=True)
        
        for row_idx, reasoning in enumerate(data['reasonings'], 2):
            ws.cell(row=row_idx, column=col_offset, value=reasoning)
        
        col_offset += 1
    
    # Format columns
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 40
    
    # Set wrap text for all columns
    for col in ws.columns:
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Create Summary sheet
    create_summary_sheet(wb, {v: d['stats'] for v, d in all_results.items()})
    
    wb.save(output_path)
    print(f"\n[scoring] Saved to {output_path}")


def score_all_versions():
    """Score all versions in UAT.xlsx."""
    configure_llm()
    
    input_path = 'C:/Users/LS0758/Downloads/UAT.xlsx'
    print(f"\n[scoring] Loading {input_path}")
    
    df = pd.read_excel(input_path)
    response_cols = get_response_columns(df)
    
    print(f"[scoring] Found {len(response_cols)} response columns")
    
    all_results = {}
    
    for col, version in response_cols:
        result = score_version(df, col, version, verbose=True)
        all_results[version] = result
    
    # Add to Excel
    add_scores_to_excel(df, all_results, input_path)
    
    return all_results


def score_latest_version():
    """Score only the latest version in UAT.xlsx."""
    configure_llm()
    
    input_path = 'C:/Users/LS0758/Downloads/UAT.xlsx'
    print(f"\n[scoring] Loading {input_path}")
    
    df = pd.read_excel(input_path)
    response_cols = get_response_columns(df)
    
    if not response_cols:
        print("[scoring] No response columns found")
        return
    
    # Get the latest (last) response column
    latest_col, latest_version = response_cols[-1]
    print(f"[scoring] Scoring latest version: {latest_version}")
    
    result = score_version(df, latest_col, latest_version, verbose=True)
    
    # Load existing results if Summary sheet exists
    all_results = {}
    try:
        wb = load_workbook(input_path)
        if 'Score Summary' in wb.sheetnames:
            summary_ws = wb['Score Summary']
            for row in summary_ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    all_results[row[0]] = {
                        'stats': {
                            'pass': row[1],
                            'partial': row[2],
                            'fail': row[3],
                            'avg': float(str(row[4]).replace('%', ''))
                        }
                    }
    except:
        pass
    
    all_results[latest_version] = result
    
    # Add to Excel
    add_scores_to_excel(df, all_results, input_path)
    
    return result


def main():
    """Main entry point."""
    if len(sys.argv) > 1 and sys.argv[1] == '--latest':
        score_latest_version()
    else:
        score_all_versions()


if __name__ == "__main__":
    main()
